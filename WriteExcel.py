import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
import os
from datetime import datetime

def creer_excel(nom_fichier, nom_projet="", localisation="", total_paye="", date_paiement="", etat_projet="", details_facturation=None):
    # Créer un nouveau classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Mes Rémunérations"

    # Obtenir le chemin absolu du dossier courant
    chemin_courant = os.path.dirname(os.path.abspath(__file__))
    logo_path = os.path.join(chemin_courant, "assets", "logo.png")

    # Styles
    titre_style = Font(name='Arial', size=24, bold=True)
    sous_titre_style = Font(name='Arial', size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Ajouter le logo si existant
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 80
        img.height = 80
        ws.add_image(img, 'A1')

    # Titre et sous-titre
    ws['C1'] = "Mes Rémunérations"
    ws['C1'].font = titre_style
    ws['C2'] = "Vous trouverez ci-dessous les rémunérations de votre projet"
    ws['C2'].font = sous_titre_style

    # Informations du projet
    info_start_row = 5
    project_info = [
        ["Nom du Projet:", nom_projet],
        ["Localisation:", localisation],
        ["Total Payé:", total_paye],
        ["Date de Paiement:", date_paiement],
        ["État du Projet:", etat_projet]
    ]

    # Style pour la boîte d'information
    for i, (label, value) in enumerate(project_info, start=info_start_row):
        ws[f'B{i}'] = label
        ws[f'C{i}'] = value
        ws[f'B{i}'].font = Font(bold=True)
        for col in ['B', 'C']:
            ws[f'{col}{i}'].border = border

    # Détails de facturation
    if details_facturation is None:
        details_facturation = [
            ["Description", "Quantité", "Prix unitaire", "Total"],
            ["Développement Frontend", "80h", "75 €/h", "6 000 €"],
            ["Développement Backend", "120h", "85 €/h", "10 200 €"],
            ["Design UI/UX", "40h", "65 €/h", "2 600 €"]
        ]

    # Titre de la section facturation
    ws['B11'] = "Détails de la Facturation"
    ws['B11'].font = Font(size=14, bold=True)

    # En-têtes du tableau
    header_row = 13
    for col, header in enumerate(details_facturation[0], start=2):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Données du tableau
    for row, data in enumerate(details_facturation[1:], start=header_row + 1):
        for col, value in enumerate(data, start=2):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    # Total général
    total_row = header_row + len(details_facturation)
    ws.cell(row=total_row + 1, column=4).value = "Total Général:"
    ws.cell(row=total_row + 1, column=5).value = total_paye
    ws.cell(row=total_row + 1, column=4).font = Font(bold=True)
    ws.cell(row=total_row + 1, column=5).font = Font(bold=True)

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Sauvegarder le fichier
    wb.save(nom_fichier)

# Exemple d'utilisation
if __name__ == "__main__":
    details_exemple = [
        ["Description", "Quantité", "Prix unitaire", "Total"],
        ["Analyse des besoins", "20h", "90 €/h", "1 800 €"],
        ["Développement", "150h", "85 €/h", "12 750 €"],
        ["Tests et déploiement", "30h", "75 €/h", "2 250 €"],
        ["Formation", "8h", "120 €/h", "960 €"]
    ]

    creer_excel(
        "mes_remunerations.xlsx",
        nom_projet="Développement Site E-commerce",
        localisation="Paris, France",
        total_paye="17 760 €",
        date_paiement="15 Mars 2024",
        etat_projet="En cours",
        details_facturation=details_exemple
    )
